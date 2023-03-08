##########################   Block 0    ##########################

from datetime import datetime
from datetime import date
import openpyxl
import pandas as pd
import sys
import os
from tkinter import messagebox
import xlsxwriter

##########################   Block 1    ##########################

## Usage function
def usage():
  print("\033[1mUsage: Script.py <text_file> <-all>-optional <days-range>-optional\033[0m\ne.g.: python script.py textfile.txt\n")
  return exit(1)

## Function to calculate age
def age(birthdate):
  today = date.today()
  age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
  return age


## Define all lists and variables
exel_file = "Birthdays.xlsx"
path = os.getcwd() + "/"
notf_before_next_bd=999
names, birthdays, ages, diff_days = ([] for i in range(4))
top_line= ["Name", "Birthday", "Age", "Days"]
today = date.today().strftime("%d/%m/%Y")

## Check arguments and valid them
if len(sys.argv) >= 2: text_file=sys.argv[1]
else: usage()
if os.path.exists(f"{path}{text_file}") == False: usage()
if len(sys.argv) >= 3 and sys.argv[2] == "-all": all=True
if len(sys.argv) == 4 and isinstance(int(sys.argv[3]), int): notf_before_next_bd = int(sys.argv[3])
else: notf_before_next_bd=999

##########################   Block 2    ##########################

## Run through file lines
with open(f'{path}{text_file}') as f:
  next(f)
  lines = [line.rstrip('\n') for line in f]

## Put Name, Birthday into lists
for line in lines:
  name = line.split(',', 1)[0]
  names.append(name)
  birthday = line.split(',', 1)[1]

  ## Save all parts of birthday into variables
  t_day = int(today[0:2])       ## today day
  t_month = int(today[3:5])     ## today month
  t_year = int(today[6:])       ## today year
  b_day = str(birthday[0:2])    ## bitrhday day
  b_month = str(birthday[3:5])  ## bitrhday month
  b_year = str(birthday[6:])    ## bitrhday year
  this_year = today[6:]         ## todays year

  ## Append Ages and Birthdays to lists
  Age = age(date(int(b_year), int(b_month), int(b_day)))
  ages.append(Age)
  birthdays.append(b_day+"/"+b_month+"/"+b_year)

  ## Calculates the differnece between today and birthday in days
  if t_month > int(b_month):
    this_year= str(int(this_year)+1)
  elif t_month == int(b_month) and t_day > int(b_day):
    this_year= str(int(this_year)+1)

  str_d1 = f"{b_day}/{b_month}/{this_year}"
  str_d2 = today

  ## Convert string to date object
  d1 = datetime.strptime(str_d1, "%d/%m/%Y")
  d2 = datetime.strptime(str_d2, "%d/%m/%Y")

  ## diff_days difference between dates in timedelta
  delta = abs(d2 - d1)
  diff_days.append(delta.days)

##########################   Block 3    ##########################

## Create exel file and give it write premission
workbook = xlsxwriter.Workbook(f"{exel_file}")
worksheet = workbook.add_worksheet()
workbook.close()
os.chmod(f"{path}{exel_file}", 0o755)

## Open the file and make changed
wb = openpyxl.load_workbook(f'{path}{exel_file}')
sheet = wb.active

## Add the headline to the exel file
for index, value in enumerate(top_line):
  sheet.cell(row=1 ,column=index+1 ,value=value)
  
def edit_exel(ROW, COL, data_list):
  for index, value in enumerate(data_list):
    sheet.cell(row=index+ROW ,column=COL ,value=value)
  
# Call the function and insert the data into columns and save
edit_exel(2, 1, names)
edit_exel(2, 2, birthdays)
edit_exel(2, 3, ages)
edit_exel(2, 4, diff_days)
wb.save(f'{path}{exel_file}')

## Sort exel file by column days until next birthday
df = pd.read_excel(f'{path}{exel_file}')
result = df.sort_values('Days', ascending = True, inplace = True)
df.to_excel(rf'{path}{exel_file}', index=False)

##########################   Block 4    ##########################

## Check the name of person with the closest birthday
days=int(min(diff_days))
name = names[diff_days.index(days)]

## Check for -all argument and days-range argument and arrange the results accordingly
msg = f"Closest birthday to {name} in {days} days"
if all == True:
  msg = ''
  temp = diff_days
  while len(temp) > 0:
    i = temp.index(min(temp))
    if temp[i] <= notf_before_next_bd:
      msg += f"{names[i]}- in {temp[i]} days\n"
    temp.pop(i)

## Output massage to screen with pop-out window
messagebox.showwarning("Birthdays", msg)
